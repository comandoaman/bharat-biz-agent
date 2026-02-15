import threading
import time
import requests
import os
import json
import openpyxl
from datetime import datetime
from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl.styles import Font
from fpdf import FPDF

# --- CONFIGURATION ---
BOT_TOKEN = "8064617293:AAHiiA8PW-Z0fBpAQKrLJp_-h9ALOlY_Ng4"
BASE_URL = f"https://api.telegram.org/bot{BOT_TOKEN}"
SERVER_PORT = 5001 
EXCEL_FILE = "store_orders.xlsx"
DATA_FILE = "nexus_state.json"
INVOICE_DIR = "invoices"

# Create invoice directory if not exists
if not os.path.exists(INVOICE_DIR):
    os.makedirs(INVOICE_DIR)

app = Flask(__name__)
CORS(app)

# Global lock
data_lock = threading.Lock()

# --- STATE MANAGEMENT ---
def load_state():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return {
        "revenue": 0,
        "orders": 0,
        "inventory": [
            {"id": 101, "name": "Cyberpunk Sneakers", "price": 4500, "stock": 12},
            {"id": 102, "name": "Neural Headphones", "price": 2500, "stock": 5},
            {"id": 103, "name": "Hologram Watch", "price": 8000, "stock": 2},
            {"id": 104, "name": "Quantum Processor", "price": 35000, "stock": 8},
        ],
        "logs": [f"[{datetime.now().strftime('%H:%M:%S')}] Nexus System Online."]
    }

db = load_state()

def save_state():
    with data_lock:
        with open(DATA_FILE, 'w') as f:
            json.dump(db, f, indent=4)

# --- 📝 HIGH-QUALITY PDF ENGINE (CORE-X EDITION) ---
class PDFInvoice(FPDF):
    def header(self):
        # -- Top Banner (Dark Grey) --
        self.set_fill_color(25, 25, 25) 
        self.rect(0, 0, 210, 40, 'F')
        
        # -- CORE-X Logo Text --
        self.set_y(10)
        self.set_font('Arial', 'B', 28)
        self.set_text_color(255, 255, 255) # White
        self.cell(10, 15, 'CORE-X', 0, 0)
        
        # -- Subtitle --
        self.set_xy(10, 22)
        self.set_font('Arial', 'B', 8)
        self.set_text_color(0, 180, 216) # Cyan Accent
        self.cell(0, 5, 'ADVANCED TECHNOLOGY SOLUTIONS', 0, 0)

        # -- "INVOICE" Label Top Right --
        self.set_xy(150, 15)
        self.set_font('Arial', 'B', 20)
        self.set_text_color(200, 200, 200)
        self.cell(50, 10, 'INVOICE', 0, 1, 'R')
        
        self.ln(20) # Spacing after banner

    def footer(self):
        self.set_y(-25)
        # -- Footer Line --
        self.set_draw_color(200, 200, 200)
        self.line(10, 275, 200, 275)
        
        # -- Footer Text --
        self.set_font('Arial', '', 8)
        self.set_text_color(100, 100, 100)
        self.cell(0, 5, 'Core-X Systems | Proprietor: Stark', 0, 1, 'C')
        self.cell(0, 5, 'Thank you for trusting the future.', 0, 1, 'C')

def generate_invoice(order_id, customer, item_name, price):
    pdf = PDFInvoice()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # --- INFO SECTION ---
    pdf.set_y(45)
    
    # Left Side: Customer Info
    pdf.set_font("Arial", 'B', 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(100, 5, "BILLED TO:", 0, 1)
    
    pdf.set_font("Arial", 'B', 14)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(100, 7, customer, 0, 1)
    
    pdf.set_font("Arial", '', 10)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(100, 5, "Telegram Authenticated User", 0, 1)
    
    # Right Side: Invoice Metadata (Moving cursor back up)
    pdf.set_xy(120, 45)
    pdf.set_font("Arial", 'B', 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(80, 5, "DETAILS:", 0, 1, 'R')
    
    pdf.set_xy(120, 52)
    pdf.set_font("Arial", '', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(80, 5, f"Date: {datetime.now().strftime('%d %b, %Y')}", 0, 1, 'R')
    
    pdf.set_xy(120, 58)
    pdf.cell(80, 5, f"Order ID: #CX-{order_id:04d}", 0, 1, 'R')

    pdf.ln(20)

    # --- TABLE SECTION ---
    # Header
    pdf.set_fill_color(240, 240, 240) # Light Grey
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(110, 10, "  Item Description", 0, 0, 'L', 1)
    pdf.cell(30, 10, "Qty", 0, 0, 'C', 1)
    pdf.cell(50, 10, "Price (USD)  ", 0, 1, 'R', 1)
    
    # Row
    pdf.set_font("Arial", '', 11)
    pdf.cell(110, 12, f"  {item_name}", 'B', 0, 'L')
    pdf.cell(30, 12, "1", 'B', 0, 'C')
    pdf.cell(50, 12, f"${price:,.2f}  ", 'B', 1, 'R')
    
    pdf.ln(5)

    # --- TOTALS ---
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(140, 10, "TOTAL PAYABLE", 0, 0, 'R')
    pdf.set_text_color(0, 180, 216) # Cyan Accent
    pdf.cell(50, 10, f"${price:,.2f}  ", 0, 1, 'R')
    
    pdf.ln(25)

    # --- SEAL & SIGNATURE (STARK) ---
    # We draw a "Digital Seal" box
    
    # 1. The Box
    x_seal = 130
    y_seal = pdf.get_y()
    pdf.set_draw_color(0, 180, 216) # Cyan
    pdf.set_line_width(0.5)
    pdf.rect(x_seal, y_seal, 60, 30)
    
    # 2. Seal Text
    pdf.set_xy(x_seal, y_seal + 2)
    pdf.set_font("Arial", 'B', 7)
    pdf.set_text_color(0, 180, 216)
    pdf.cell(60, 5, "DIGITALLY AUTHORIZED", 0, 1, 'C')
    
    # 3. Signature (Simulated with Italic Serif)
    pdf.set_xy(x_seal, y_seal + 8)
    pdf.set_font("Times", 'BI', 16) # Times Bold Italic looks like handwriting
    pdf.set_text_color(0, 0, 0)
    pdf.cell(60, 10, "Stark", 0, 1, 'C')
    
    # 4. Seal Footer
    pdf.set_xy(x_seal, y_seal + 22)
    pdf.set_font("Arial", '', 6)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(60, 5, "CORE-X OFFICIAL SEAL", 0, 1, 'C')

    # Output
    filename = f"{INVOICE_DIR}/CoreX_Invoice_{order_id}.pdf"
    pdf.output(filename)
    return filename

# --- 🛠️ CORE UTILITIES (UNCHANGED) ---
def log_event(message):
    timestamp = datetime.now().strftime("%H:%M:%S")
    entry = f"[{timestamp}] {message}"
    with data_lock:
        db["logs"].insert(0, entry)
        db["logs"] = db["logs"][:50] 
    print(entry)

def save_to_excel(customer, item, price):
    try:
        current_order_id = db["orders"] + 1
        with data_lock:
            if not os.path.exists(EXCEL_FILE):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Sales"
                headers = ["Order ID", "Customer", "Item", "Price", "Time"]
                ws.append(headers)
                for cell in ws[1]: cell.font = Font(bold=True)
            else:
                wb = openpyxl.load_workbook(EXCEL_FILE)
                ws = wb.active
            
            current_order_id = ws.max_row 
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([current_order_id, customer, item, price, timestamp])
            wb.save(EXCEL_FILE)
            return current_order_id
    except Exception as e:
        log_event(f"❌ EXCEL ERROR: {str(e)}")
        return None

# --- 🤖 TELEGRAM LOGIC (UNCHANGED) ---
def send_tg_message(chat_id, text, reply_markup=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    requests.post(f"{BASE_URL}/sendMessage", json=payload)

def send_tg_document(chat_id, file_path):
    url = f"{BASE_URL}/sendDocument"
    try:
        with open(file_path, 'rb') as f:
            files = {'document': f}
            data = {'chat_id': chat_id, 'caption': '📄 Here is your CORE-X receipt.'}
            requests.post(url, data=data, files=files)
    except Exception as e:
        log_event(f"Failed to send PDF: {e}")

def get_stock_keyboard():
    buttons = []
    for item in db["inventory"]:
        status = "✅" if item["stock"] > 0 else "❌"
        buttons.append([{"text": f"{status} {item['name']} - ${item['price']}", "callback_data": f"buy_{item['id']}"}])
    return {"inline_keyboard": buttons}

def handle_bot_logic(user, chat_id, text):
    text = text.lower()
    
    if "/start" in text or "help" in text:
        msg = "🚀 *Welcome to CORE-X*\nClick a button below to purchase or use `/stock` to see items."
        send_tg_message(chat_id, msg, reply_markup=get_stock_keyboard())
        return

    if text == "/stock":
        send_tg_message(chat_id, "*📦 CORE-X Inventory:*", reply_markup=get_stock_keyboard())
        return

    if text == "/stats":
        revenue = db['revenue']
        orders = db['orders']
        msg = f"📊 *CORE-X Statistics*\n\n💰 Revenue: ${revenue}\n📦 Total Orders: {orders}\n"
        send_tg_message(chat_id, msg)
        return

    matched_item = None
    for item in db["inventory"]:
        if item["name"].lower() in text:
            matched_item = item
            break
    
    if matched_item:
        process_purchase(user, chat_id, matched_item)
    else:
        send_tg_message(chat_id, "❓ Unknown command. Try `/stock` or type an item name.")

def process_purchase(user, chat_id, item):
    if item["stock"] > 0:
        with data_lock:
            item["stock"] -= 1
            db["revenue"] += item["price"]
            db["orders"] += 1
        
        order_id = save_to_excel(user, item['name'], item['price'])
        save_state()
        
        log_event(f"SALE: {user} bought {item['name']}")
        
        confirm_msg = f"✅ *Purchase Successful!*\n\n*Item:* {item['name']}\n*Price:* ${item['price']}\n*Stock Remaining:* {item['stock']}\n\n_Generating CORE-X Invoice..._"
        send_tg_message(chat_id, confirm_msg)

        if order_id:
            def invoice_task():
                pdf_path = generate_invoice(order_id, user, item['name'], item['price'])
                send_tg_document(chat_id, pdf_path)
            
            threading.Thread(target=invoice_task).start()
            
    else:
        send_tg_message(chat_id, f"❌ *Out of Stock:* {item['name']} is currently unavailable.")

def handle_callback(update):
    query = update["callback_query"]
    chat_id = query["message"]["chat"]["id"]
    user = query["from"].get("first_name", "User")
    data = query["data"]

    try:
        requests.post(f"{BASE_URL}/answerCallbackQuery", json={"callback_query_id": query["id"]})
    except:
        pass

    if data.startswith("buy_"):
        item_id = int(data.split("_")[1])
        item = next((i for i in db["inventory"] if i["id"] == item_id), None)
        if item:
            process_purchase(user, chat_id, item)

# --- 🛰️ NETWORK THREADS (UNCHANGED) ---
def bot_polling_loop():
    offset = 0
    requests.post(f"{BASE_URL}/deleteWebhook")
    print("🛰️ Telegram Polling Started...")
    
    while True:
        try:
            res = requests.get(f"{BASE_URL}/getUpdates", params={"offset": offset, "timeout": 20}).json()
            if "result" in res:
                for update in res["result"]:
                    offset = update["update_id"] + 1
                    
                    if "message" in update and "text" in update["message"]:
                        msg = update["message"]
                        handle_bot_logic(msg["from"].get("first_name", "User"), msg["chat"]["id"], msg["text"])
                    
                    elif "callback_query" in update:
                        handle_callback(update)
                        
        except Exception as e:
            print(f"Polling Error: {e}")
            time.sleep(5)
        time.sleep(0.5)

# --- 🌐 FLASK API (UNCHANGED) ---
@app.route('/data', methods=['GET'])
def get_data():
    return jsonify(db)

@app.route('/update-inventory', methods=['POST'])
def update_inventory():
    data = request.json
    item_id = data.get("id")
    new_stock = data.get("stock")
    
    with data_lock:
        for item in db["inventory"]:
            if item["id"] == item_id:
                item["stock"] = new_stock
                save_state()
                log_event(f"ADMIN: Updated {item['name']} stock to {new_stock}")
                return jsonify({"status": "success"})
    return jsonify({"status": "error", "message": "Item not found"}), 404

@app.route('/reset-logs', methods=['POST'])
def reset_logs():
    with data_lock:
        db["logs"] = [f"[{datetime.now().strftime('%H:%M:%S')}] Logs Cleared by Admin."]
        save_state()
    return jsonify({"status": "success"})

# --- 🚀 LAUNCHPAD ---
if __name__ == '__main__':
    threading.Thread(target=bot_polling_loop, daemon=True).start()
    print(f"🚀 CORE-X BACKEND ONLINE: http://127.0.0.1:{SERVER_PORT}")
    print(f"📂 Invoice Folder: {os.path.abspath(INVOICE_DIR)}")
    app.run(host='0.0.0.0', port=SERVER_PORT, debug=False, use_reloader=False)